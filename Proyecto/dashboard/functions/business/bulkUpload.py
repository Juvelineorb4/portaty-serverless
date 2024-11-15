import boto3
import json
import openpyxl
import os
import uuid
import re
import requests
from botocore.exceptions import ClientError
from openpyxl import Workbook
from datetime import datetime

# Obtener el nombre del bucket desde las variables de entorno
bucket_name = os.environ.get("S3_BUCKET_NAME")
PLACE_INDEX = os.environ.get("PLACE_INDEX_NAME")
AWS_REGION = os.environ.get("AWS_REGION")
TABLE_BUSINESS_NAME = os.environ.get("TABLE_BUSINESS")
APPSYNC_API_URL = os.environ.get("APPSYNC_URL")
APPSYNC_API_KEY = os.environ.get("APPSYNC_API_KEY")
# Cliente de S3
s3 = boto3.client('s3')
# CLiente location
place = boto3.client('location', region_name=AWS_REGION)
#CLiente Dynamodb
# Cambiar la instancia de DynamoDB a un recurso en lugar de un cliente
dynamodb = boto3.client("dynamodb", region_name=AWS_REGION)


mutation = """
    mutation CreateBusinessUploadHistory(
    $input: CreateBusinessUploadHistoryInput!
    $condition: ModelBusinessUploadHistoryConditionInput
  ) {
    createBusinessUploadHistory(input: $input, condition: $condition) {
      id
      uploadDate
      completionDate
      validCount
      total
      invalidCount
      originalPath
      invalidPath
      createdAt
      updatedAt
      __typename
    }
  }
    """ 


# Función Lambda principal con ajustes para los campos específicos
def handler(event, context):
    try:
        # Obtener el nombre del bucket y la clave del archivo desde el evento
        key = event['Records'][0]['s3']['object']['key']
        
        # Descargar el archivo desde S3
        local_file_path = '/tmp/' + key.split('/')[-1]
        s3.download_file(bucket_name, key, local_file_path)

        # Cargar el archivo XLSX
        workbook = openpyxl.load_workbook(local_file_path)
        sheet = workbook.active
        
        # Validar y procesar cada fila
        valid_rows = []
        invalid_rows = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            print(f"Procesando fila {row_index}: {row}")
            
            # Obtener los campos básicos
            id = row[0]  # La columna de 'id'
            name = row[1]  # La columna de 'name'
            location = row[2]  # La columna de 'location'
            area = row[3]  # La columna de 'area'
            activity = row[4]  # La columna de 'activity'

            # Inicializar valores para phone, email e image
            phone = ""
            email = ""
            image = ""
            
        
            # Validar la columna 5 (phone, email o image)
            if len(row) > 5:
                print("ROW 5",row[5] )
                value = row[5]  # Obtener el valor de la columna 5
                if validate_phone(value):  # Si es un teléfono válido
                    phone = value
                elif validate_email(value):  # Si es un email válido
                    email = value
                else:  # Si no es ni teléfono ni email, asumimos que es una imagen
                    image = value

            # Validar la columna 6 (si hay más valores y no se llenaron ya en phone o email)
            if len(row) > 6 and not email:  # Si email aún no está lleno, validamos la columna 6
                print("ROW 6",row[6])
                value = row[6]
                if validate_email(value):  # Si es un email válido
                    email = value
                else:  # Si no es ni teléfono ni email, asumimos que es una imagen
                    image = value

            # Generar un nuevo ID usando uuid
            new_id = str(uuid.uuid4())

            # Validar cada campo y agregar un motivo si falla
            if not isinstance(name, str) or not name.strip():
                print(f"Fila {row_index}: 'name' inválido: {name}")
                invalid_rows.append({"row": row, "motivo": "'name' inválido o vacío"})
                continue
            if not isinstance(area, str) or not area.strip():
                print(f"Fila {row_index}: 'area' inválido: {area}")
                invalid_rows.append({"row": row, "motivo": "'area' inválido o vacío"})
                continue
            if not isinstance(activity, str) or not activity.strip():
                print(f"Fila {row_index}: 'activity' inválido: {activity}")
                invalid_rows.append({"row": row, "motivo": "'activity' inválido o vacío"})
                continue
            if not validate_phone(phone):
                print(f"Fila {row_index}: 'phone' inválido: {phone}")
                invalid_rows.append({"row": row, "motivo": "'phone' no válido"})
                continue
            if not validate_email(email):
                print(f"Fila {row_index}: 'email' inválido: {email}")
                invalid_rows.append({"row": row, "motivo": "'email' no válido"})
                continue

  


            # Validar la imagen y subirla a S3, obteniendo la URL si es válida
            image_s3_url = validate_and_upload_image(image, bucket_name, new_id)
            if not image_s3_url:
                print(f"Fila {row_index}: 'image' inválida o no se pudo cargar: {image}")
                invalid_rows.append({"row": row, "motivo": "'image' no válida o no se pudo cargar"})
                continue
            
            if not validate_location(location):
                print(f"Fila {row_index}: 'location' inválido: {location}")
                invalid_rows.append({"row": row, "motivo": "'location' no válido o incompleto"})
                continue
            
            # Convertir la location a JSON y obtener la dirección
            location_json = json.loads(location)
            address = get_address_from_location(location_json['latitude'], location_json['longitude'])
            address_string = json.dumps(address)
            if not address:
                print(f"Fila {row_index}: 'location' no pudo ser resuelta a una dirección")
                invalid_rows.append({"row": row, "motivo": "'location' no pudo ser resuelta a una dirección"})
                continue
            
            # Crear el objeto de coordenadas en formato de mapa (M)
            coordinates_map = {
                "lat": {"N": str(location_json["latitude"])},
                "lon": {"N": str(location_json["longitude"])}
            }
             # Obtener el timestamp actual en formato ISO 8601
            timestamp = get_iso_timestamp()

            # Construir los tags como lista de cadenas JSON serializadas
            # Construir los tags como una lista de strings
            tags = [
                json.dumps({"priority": 1, "value": name, "date": timestamp}),
                json.dumps({"priority": 2, "value": area, "date": timestamp}),
                json.dumps({"priority": 2, "value": activity, "date": timestamp}),
                json.dumps({"priority": 1, "value": address_string, "date": timestamp})
            ]
            images = [json.dumps({"key": 0, "url": image_s3_url, "description":"", "date":timestamp})]

            # Construir el objeto para DynamoDB
            # Construir el objeto para DynamoDB
            business_item = {
                "id": {"S": new_id},
                "name": {"S": name},
                "thumbnail": {"S": image_s3_url},
                "email": {"S": email or ""},  # Asignar cadena vacía si email es None
                "phone": {"S": phone or ""},  # Asignar cadena vacía si phone es None
                "coordinates": {"M": coordinates_map},
                "activity": {"S": json.dumps({"main": area, "sub": activity})},
                "tags": {"L": [{"S": tag} for tag in tags]},  # Convertir cada elemento a un string en DynamoDB
                "images": {"L":[{"S": image} for image in images]},
                "status": {"S": "ENABLED"},
                "statusOwner": {"S": "NOT_ASSIGNED"},
                "createdAt": {"S":timestamp},
                "updatedAt": {"S":timestamp},
                "__typename": {"S": "Business"}
            }


               

            # Agregar la fila validada
            valid_rows.append(business_item)

        # Si hay filas válidas, subirlas a DynamoDB en formato batch
        if valid_rows:
            upload_to_dynamodb(valid_rows)

        # Si hay filas inválidas, guardar y subir el archivo
        if invalid_rows:
            invalid_file_path = '/tmp/invalid_business.xlsx'
            save_invalid_rows(invalid_rows, invalid_file_path)
            invalid_key =  upload_invalid_business_file(bucket_name, invalid_file_path)

        # Rutas de los archivos en S3
        invalid_file_s3_path = f"https://{bucket_name}.s3.amazonaws.com/{invalid_key}" if invalid_rows else None
        original_file_s3_path = f"https://{bucket_name}.s3.amazonaws.com/{key}"
        

        # Obtener el timestamp de finalización
        completion_timestamp = get_iso_timestamp()

        # Crear el historial en DynamoDB usando AppSync
        history_params = {
            "uploadDate": completion_timestamp,
            "completionDate": completion_timestamp,
            "validCount": len(valid_rows),
            "invalidCount": len(invalid_rows),
            "total": len(valid_rows) + len(invalid_rows),
            "originalPath": original_file_s3_path,
            "invalidPath": invalid_file_s3_path if invalid_file_s3_path else ""
        }
        create_history(history_params)  # Crear el historial de la subida

        print("valid_rows:", valid_rows)
        print("invalid_rows:", invalid_rows)
        print(f"CARGA DE NEGOCIOS MASIVOS EXITOSA: {len(valid_rows)} registros válidos")
        return {
            "message": "Carga de negocio exitosa.",
            "valid_records": len(valid_rows),
            "invalid_records": len(invalid_rows)
        }
    except ClientError as e:
        print(f"Error al acceder a S3: {e}")
        return {"error": str(e)}
    except Exception as e:
        print(f"Error en carga de búsqueda de negocios: {str(e)}")
        return {"error": str(e)}
    


def create_history(params):
    # Construir el cuerpo de la solicitud (variables y headers)
    variables = {
        "input": {
            "uploadDate": params["uploadDate"],
            "completionDate": params["completionDate"],
            "validCount": params["validCount"],
            "total": params["total"],
            "invalidCount": params["invalidCount"],
            "originalPath": params["originalPath"],
            "invalidPath": params["invalidPath"]
        }
    }

    headers = {
        "Content-Type": "application/json",
        "x-api-key": APPSYNC_API_KEY
    }

    # Enviar la solicitud POST a AppSync
    try:
        response = requests.post(
            APPSYNC_API_URL,
            json={"query": mutation, "variables": variables},
            headers=headers
        )
        response_data = response.json()

        # Verificar si la respuesta contiene errores
        if "errors" in response_data:
            print("Error en la mutación de AppSync:", response_data["errors"])
            return {"error": response_data["errors"]}

        print("Historial de carga creado exitosamente:", response_data)
        return response_data

    except Exception as e:
        print(f"Error al crear el historial en AppSync: {str(e)}")
        return {"error": str(e)}
# Función para generar una marca de tiempo con formato ISO 8601
def get_iso_timestamp():
    # Obtener el tiempo actual en UTC y formatearlo en ISO 8601 con milisegundos
    return datetime.utcnow().isoformat(timespec='milliseconds') + 'Z'
# Función para subir los elementos a DynamoDB
def upload_to_dynamodb(items):
    try:
        print(TABLE_BUSINESS_NAME)
        # Divide los items en lotes de 25 (límite de DynamoDB para operaciones de batch)
        batches = [items[i:i + 25] for i in range(0, len(items), 25)]
        
        for batch in batches:
            request_items = {TABLE_BUSINESS_NAME: []}
            
            for item in batch:
                request_items[TABLE_BUSINESS_NAME].append({
                    'PutRequest': {
                        'Item': item
                    }
                })
            
            if not request_items[TABLE_BUSINESS_NAME]:
                print("No hay elementos en el lote para cargar.")
                continue
            
            # Llamada a batch_write_item para cargar el lote en DynamoDB
            response = dynamodb.batch_write_item(RequestItems=request_items)
            print(f"Lote de {len(batch)} elementos subido exitosamente a DynamoDB. Respuesta: {response}")
    except ClientError as e:
        print(f"Error al cargar elementos en DynamoDB: {e}")
    except Exception as e:
        print(f"Error inesperado al cargar en DynamoDB: {e}")





# Función para obtener la dirección a partir de las coordenadas
def get_address_from_location(latitude, longitude):
    try:
        response = place.search_place_index_for_position(
            IndexName=PLACE_INDEX,
            Position=[longitude, latitude],
            MaxResults=1
        )
        if response['Results']:
            place_data = response['Results'][0]['Place']
            address = {
                "city": place_data.get("Municipality", None),
                "timezone": place_data.get("TimeZone", None),
                "subregion": place_data.get("SubRegion", None),
                "formattedAddress": place_data.get("Label", None),
                "isoCountryCode": place_data.get("CountryCode", None),
                "country": place_data.get("Country", None),
                "district": place_data.get("Neighborhood", None),
                "name": place_data.get("AddressNumber", None),
                "region": place_data.get("Region", None),
                "streetNumber": place_data.get("AddressNumber", None),
                "postalCode": place_data.get("PostalCode", None),
                "street": place_data.get("Street", None)
            }
            return address
        else:
            print(f"No se encontró una dirección para las coordenadas: {latitude}, {longitude}")
            return None
    except ClientError as e:
        print(f"Error al obtener la dirección desde AWS Location: {e}")
        return None
    except Exception as e:
        print(f"Error inesperado al obtener la dirección: {e}")
        return None
    
def save_invalid_rows(invalid_rows, file_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['id','Name', 'Location', 'Area', 'Activity', 'Phone', 'Email', 'Image', 'Motivo'])
    for entry in invalid_rows:
        row = entry["row"]
        motivo = entry["motivo"]
        sheet.append(list(row) + [motivo])
    workbook.save(file_path)

def upload_invalid_business_file(bucket_name, file_path):
    try:
        # Obtener la fecha y hora actual para estructurar el path
        now = datetime.now()
        year = now.year
        month = f"{now.month:02d}"
        day = f"{now.day:02d}"
        hour = f"{now.hour:02d}"
        minute = f"{now.minute:02d}"
        second = f"{now.second:02d}"
        
        # Definir el path en S3
        s3_key = f"public/document/invalid/business/{year}-{month}-{day}/{hour}-{minute}-{second}/invalid_business.xlsx"
        print("KEY DE FILE XLSX: ", s3_key)
        s3.upload_file(file_path, bucket_name, s3_key)
        print(f"Archivo {file_path} subido exitosamente a s3://{bucket_name}/{s3_key}")
        return s3_key
    except ClientError as e:
        print(f"Error subiendo el archivo de negocios inválidos a S3: {e}")

# Función para validar y cargar una imagen en S3
def validate_and_upload_image(url, bucket_name, business_id):
    try:
        response = requests.get(url, stream=True)
        if response.status_code == 200:
            content_type = response.headers.get('Content-Type')
            if 'image' in content_type:
                file_extension = content_type.split('/')[1]
                local_image_path = f"/tmp/{business_id}.{file_extension}"
                with open(local_image_path, 'wb') as img_file:
                    for chunk in response.iter_content(1024):
                        img_file.write(chunk)
                
                s3_key = f"public/assets/business/{business_id}/image.{file_extension}"
                s3_url = upload_to_s3(bucket_name, s3_key, local_image_path)
                return s3_url
            else:
                print(f"El contenido no es una imagen: {content_type}")
                return None
        else:
            print(f"No se pudo acceder a la imagen en la URL: {url}, status code: {response.status_code}")
            return None
    except Exception as e:
        print(f"Error al validar o cargar la imagen desde la URL {url}: {e}")
        return None

def upload_to_s3(bucket_name, s3_key, file_path):
    try:
        s3.upload_file(file_path, bucket_name, s3_key)
        print(f"Archivo {file_path} subido exitosamente a s3://{bucket_name}/{s3_key}")
        s3_url = f"https://{bucket_name}.s3.amazonaws.com/{s3_key}"
        return s3_url
    except ClientError as e:
        print(f"Error subiendo el archivo a S3: {e}")
        return None

# Función para validar el teléfono
def validate_phone(phone):
    
    if not phone:  # Permitir valor vacío
        print("TELEFONO VACIO")
        return True
    
    pattern = r'^\+\d{1,3}\d{4,14}$'
    if not re.match(pattern, phone):
        print(f"Teléfono no válido: {phone}")
        return False
    print("TELEFONO VALIDO")
    return True

# Función para validar el email
def validate_email(email):
    if not email:  # Permitir valor vacío
        print("EMAIL VACIO")
        return True
    if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
        print(f"Email no válido: {email}")
        return False
    print("EMAIL VALIDO")
    return True

# Función para validar el campo de location
def validate_location(location):
    if isinstance(location, str):
        try:
            location = json.loads(location)
        except json.JSONDecodeError:
            print(f"Location no es un JSON válido: {location}")
            return False
    if not isinstance(location, dict):
        print(f"Location no es un diccionario: {location}")
        return False
    if 'latitude' not in location or 'longitude' not in location:
        print(f"Faltan campos en location: {location}")
        return False
    if not isinstance(location['latitude'], (int, float)) or not isinstance(location['longitude'], (int, float)):
        print(f"Valores inválidos en location: {location}")
        return False
    return True
